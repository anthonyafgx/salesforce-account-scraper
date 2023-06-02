from types import NoneType
from Scraper import Scraper
from enum import Enum
import openpyxl
import threading
import shutil

class Columns(Enum):
    NAME = 1
    PHONE = 2
    STREET = 3
    CITY = 4
    STATE = 5
    ZIP = 5
    COUNTRY = 6

# Excel read and write operations
# Controls Scraper's mutlithreading
class ExcelController():
    def __init__(self, 
                excel_path, 
                name_col, 
                phone_col, 
                street_col, 
                city_col,
                state_col,
                zip_col,
                country_col, 
                num_of_threads=4):

        # identifies the position of the column
        self.columns = {Columns.NAME: name_col,
                        Columns.PHONE: phone_col,
                        Columns.STREET: street_col,
                        Columns.CITY: city_col,
                        Columns.STATE: state_col,
                        Columns.ZIP: zip_col,
                        Columns.COUNTRY: country_col}

        # excel path
        self.excel_path = excel_path

        # multithreading variables
        self.num_of_threads = num_of_threads
        self.account_name = None

        # open excel file and sheet
        # the sheet is used to edit, the file is used to save
        self.excel_file = openpyxl.load_workbook(excel_path, read_only=True)
        self.sheet = self.excel_file.active

        # create a driver for each tread
        self.drivers = [Scraper() for x in range(num_of_threads)]

        # create a lock for all threads
        self.thread_lock = threading.Lock()

        # account names
        self.account_names = self._get_account_names()
        self.current_row = 1

        # accounts data
        # is a list of dictionaries that have the following structure
        # ret_dict = {'shipping-street': "56th St.", 
        #             'shipping-city': "New York City", 
        #             'shipping-state': "NY", 
        #             'shipping-zip': "33126",
        #             'shipping-country': "United States", 
        #             'website': "65.com", 
        #             'phone': "888-999-1010"}
        self.accounts_data = []

    # retrieves the account names from the excel sheet
    # returns a list with all the account names
    def _get_account_names(self):
        account_names = []
        for row in range(2, 100):   # <--- NOTE: CHANGE THIS HARDCODED "100" TO THE MAX ROW 
            account_names.append(self.sheet.cell(row=row, column=self.columns[Columns.NAME]).value)
        return account_names

    # main funciton of this object
    # scrapes the whole excel
    def scrape_excel(self):
        # Scrape all the info with multithreading
        while len(self.account_names) != 0:
            self._thread_main_task()

        # Close the original excel file (which was opened as read-only for optimization)
        self.excel_file.close()

        # Create a copy of the excel file
        # output_path = self.excel_path + ' OUTPUT.xlsx'
        # shutil.copyfile(self.excel_path, output_path)

        # Write all the info to a new excel file
        output_excel = openpyxl.Workbook(write_only=False)
        output_sheet = output_excel.create_sheet()
        def write_helper(row, col, val):
            output_sheet.cell(row = row, column = col).value = val
            # if output_sheet.cell(row = row, column = col).value == None:
            #     output_sheet.cell(row = row, column = col).value = val
            # else:
            #     print("[LOG] Won't overwrite. Value already written.")
        
        # header
        output_sheet.cell(row=1, column=1, value="Account Name")
        output_sheet.cell(row=1, column=2, value="Website")
        output_sheet.cell(row=1, column=3, value="Phone")
        output_sheet.cell(row=1, column=4, value="Shipping Street")
        output_sheet.cell(row=1, column=5, value="Shipping City")
        output_sheet.cell(row=1, column=6, value="Shipping State")
        output_sheet.cell(row=1, column=7, value="Shipping ZIP")
        output_sheet.cell(row=1, column=8, value="Shipping Country")
        for account in self.accounts_data:
            write_helper(account['row'],1,account['name'])
            write_helper(account['row'],2,account['website'])
            write_helper(account['row'],3,account['phone'])
            write_helper(account['row'],4,account['shipping-street'])
            write_helper(account['row'],5,account['shipping-city'])
            write_helper(account['row'],6,account['shipping-state'])
            write_helper(account['row'],7,account['shipping-zip'])
            write_helper(account['row'],8,account['shipping-country'])

        # save
        output_excel.save(self.excel_path + ' OUTPUT.xlsx')

    # the task each thread points to
    def _thread_scrape(self, driver):
        self.thread_lock.acquire()  # lock
        try: # if not empty
            account_name = self.account_names.pop() # get last element and eliminate from account names
            self.current_row += 1
            local_row = self.current_row
        except IndexError:
            self.thread_lock.release()  # release
            return None
        self.thread_lock.release()      # release
        # scrape the information
        account_data = driver.scrape(account_name)
        # add name and row to account's data
        account_data['name'] = account_name      # add name information to dictionary
        account_data['row'] = local_row   # add row information to dictionary
        # add data dict to main list
        self.thread_lock.acquire()
        self.accounts_data.append(account_data)  # scrape and add to main data dictionary
        self.thread_lock.release()

    def _thread_main_task(self):
        # create a thread for each driver
        self.threads = [threading.Thread(target=self._thread_scrape, args=(driver,)) for driver in self.drivers]

        # start all threads
        for t in self.threads:
            t.start()

        # wait for threads to end the function
        for t in self.threads:
            t.join()
        
e = ExcelController('test_data.xlsx',
                    3,4,5,6,7,8,9,8)

e.scrape_excel()
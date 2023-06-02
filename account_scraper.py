from concurrent.futures import thread
from csv import excel
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.chart import BarChart, Reference
import string
import sys
import os

from Scraper import Scraper

excelPath = 'Account Data Clean-up 6.24.22.xlsx'
excelFile = openpyxl.load_workbook(excelPath)
sheet = excelFile.active
 
# Create Scraper
s = None    # assigned below

# Arguments
range_start = None
range_end = None

def write(row_number, phone, website, street, city, state, zip, country): 
    # columnas de 4 al 10
    # llenar las celdas del excel   
    def write_helper(col, var):
        if sheet.cell(row = row_number, column = col).value == None:
            sheet.cell(row = row_number, column = col).value = var
        else:
            s._log_success("[LOG] Won't overwrite. Value already written.")
    
    write_helper(4,phone)
    write_helper(5,website)
    write_helper(7,street)
    write_helper(8,city)
    write_helper(9,state)
    write_helper(10,zip)
    write_helper(11,country)

    excelFile.save(excelPath)

def main():


    max_row = sheet.max_row + 1

    for i in range(range_start, range_end + 1):
        # get account name
        current_account = sheet.cell(row = i, column = 3).value
        
        # scrape data
        data_dict = s.scrape(current_account)

        # insert data
        write(i, data_dict['phone'],
                data_dict['website'],
                data_dict['shipping-street'],
                data_dict['shipping-city'],
                data_dict['shipping-state'],
                data_dict['shipping-zip'],
                data_dict['shipping-country'])

        # progress counter
        print(str(i) + "/" + str(max_row))

        
    

if __name__ == '__main__':
    # Arguments processing
    if len(sys.argv) <= 1:
        print("[ERROR] You must provide a range of rows to process. Example: file.py 10 20")
        os.system("pause")
        sys.exit()
    else:
        try:
            range_start = int(sys.argv[1])
            range_end = int(sys.argv[2])
        except:
            print("[ERROR] Some error happened parsing the arguments")
        else:
            s = Scraper()
            s._log_success("[SETUP] Arguments parsed correctly. The program will process rows from row " + str(range_start) + " to row " + str(range_end))
            main()
